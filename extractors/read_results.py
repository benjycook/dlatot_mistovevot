# -*- coding: utf-8 -*-


import os
from openpyxl import load_workbook
import re


def main():
    results_path = os.getcwd()+'/results'
    files = [f for f in os.listdir(results_path)] # if f.startswith('app')]
    c = 0
    ra = 0
    for f in files:
        wb = load_workbook('{}/{}'.format(results_path, f)) #, read_only=True)
        # split = re.split(r'_|\.', f)
        # if f.startswith('app'):
        #     year = split[1]
        #     ws = wb[year]
        #     c += ws.max_row
        # else:
        #     year = split[2]
        #     ws = wb[year]
        #     ra += ws.max_row
        # std = wb.get_sheet_by_name('Sheet')
        # wb.remove(std)
        # wb.save('{}/{}'.format(results_path, f))
        # print(ws.max_row)
        print(wb.get_sheet_names())
    print(c)
    print(ra)

if __name__ == '__main__':
    main()

