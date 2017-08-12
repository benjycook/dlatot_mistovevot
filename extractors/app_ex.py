# -*- coding: utf-8 -*-

from bs4 import BeautifulSoup
import pandas as pd
import os
import requests
from time import sleep
from tqdm import tqdm
from openpyxl import Workbook

from tofes_class import Tofes


def get_csv_data(argv):
    action = argv[1]
    _year = argv[2]
    filename = os.path.dirname(os.getcwd())+'{}{}_{}.{}'.format('/data/scraped_data/', action, _year, 'csv')
    df = pd.read_csv(filename)
    return df


def fill_review_again_sheet(sheet, row, info, failure_reason):
    a, b, c, d, e = ['{}{}'.format(c, row) for c in ['A', 'B', 'C', 'D', 'E']]
    sheet[a] = info.action
    sheet[b] = info.company_name
    sheet[c] = info.maya_link
    sheet[d] = info.html_link
    sheet[e] = failure_reason


def main():

    for x, year in enumerate(range(2004, 2018)):
        wb = Workbook()
        wb_ra = Workbook()
        year = str(year)
        ws = wb.create_sheet(year, x)
        ws['A1'] = u'שם מלא'
        ws['B1'] = u'תפקיד'
        ws['C1'] = u'תאור תפקיד'
        ws['D1'] = u'תחילת תפקיד'
        ws['E1'] = u'שם החברה'
        ws['F1'] = u'משרות קודמות'
        ws['G1'] = u'תאור פעולה'
        ws['H1'] = u'מספר טופס'
        ws['I1'] = u'תאריך פרסום'
        ws['J1'] = u'לינק רפרנס'

        ws_ra = wb_ra.create_sheet(year, x)
        ws_ra['A1'] = 'action'
        ws_ra['B1'] = 'company_name'
        ws_ra['C1'] = 'maya_link'
        ws_ra['D1'] = 'tofes_link'
        ws_ra['E1'] = 'failure_reason'

        # print('\n{}'.format(year))
        r = 1
        ra = 1
        for i, data in tqdm(enumerate(get_csv_data(['', 'appointment', year]).iterrows())):
            sleep(0.5)
            # if i > 3:
            #     break
            item = data[1]
            tofes_link = item.html_link
            item_info = {'year': year,
                         'action': item.action,
                         'company_name': item.company_name,
                         'maya_link': item.maya_link,
                         'tofes_link': tofes_link}

            if tofes_link == 'nan' or type(tofes_link) != str or len(tofes_link) < 10 or tofes_link.endswith('.pdf'):
                ra += 1
                fill_review_again_sheet(ws_ra, ra, item, 'bad tofes link')
                continue
            try:
                tofes = Tofes(item_info)
                if not tofes.status:
                    ra += 1
                    fill_review_again_sheet(ws_ra, ra, item, 'request failed')
                    continue

                if not tofes.valid_report_num:
                    ra += 1
                    fill_review_again_sheet(ws_ra, ra, item, u'bad tofes num - {}'.format(tofes.report_num))
                    continue

                r += 1
                a, b, c, d, e, f, g, h, i, j = ['{}{}'.format(c, r)
                                                for c in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']]
                ws[a] = tofes.fullname
                ws[b] = tofes.job_title
                ws[c] = tofes.job_desc
                ws[d] = tofes.starting_date
                ws[e] = item.company_name
                ws[f] = ','.join(tofes.prior_jobs)
                ws[g] = item.action
                ws[h] = tofes.report_num
                ws[i] = tofes.date_published
                ws[j] = tofes_link
            except:
                ra += 1
                fill_review_again_sheet(ws_ra, ra, item, u'new exception')
                print(tofes_link)

        wb.save('results/appointments_{}.xlsx'.format(year))
        wb_ra.save('results/review_again_{}.xlsx'.format(year))


if __name__ == '__main__':
    main()

